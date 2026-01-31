"""
Report Generator for AdInsights-Agent

Generates professional, stakeholder-ready reports with:
- Markdown formatting
- Embedded charts
- Executive summaries
- Actionable recommendations
- Data tables

Export formats:
- Markdown (.md)
- HTML (.html)
- PDF (.pdf) - optional, requires weasyprint
"""

import os
import re
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

import pandas as pd
import numpy as np

# Try to import optional dependencies
try:
    import markdown
    from markdown.extensions import tables, fenced_code, nl2br
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# REPORT GENERATOR CLASS
# =============================================================================

class ReportGenerator:
    """
    Generate professional analytics reports.

    Features:
    - Markdown generation with embedded charts
    - Executive summary creation
    - Actionable recommendations
    - HTML and PDF export
    - Professional styling
    """

    def __init__(
        self,
        output_dir: str = "./data/reports",
        company_name: str = "AdInsights Analytics",
        include_timestamps: bool = True,
        include_raw_data: bool = False,
    ):
        """
        Initialize the report generator.

        Args:
            output_dir: Directory to save reports
            company_name: Company name for header
            include_timestamps: Include timestamps in reports
            include_raw_data: Include raw data tables in appendix
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.company_name = company_name
        self.include_timestamps = include_timestamps
        self.include_raw_data = include_raw_data

        # CSS for HTML/PDF styling
        self._css = self._get_default_css()

    # =========================================================================
    # REPORT SECTION GENERATORS
    # =========================================================================

    def create_executive_summary(
        self,
        insights: List[str],
        metrics: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Create a concise executive summary (2-3 sentences).

        Format:
        - Overview sentence: What was analyzed
        - Key finding: Most important insight
        - Action item: Top priority recommendation

        Args:
            insights: List of insight strings
            metrics: Optional metrics dictionary for context

        Returns:
            Executive summary as markdown string
        """
        lines = []

        # Overview
        if metrics:
            campaign_id = metrics.get("campaign_id", "Unknown")
            total_days = metrics.get("total_days", metrics.get("record_count", "N/A"))
            total_spend = metrics.get("total_spend", metrics.get("avg_spend", 0))

            if isinstance(total_spend, (int, float)):
                spend_str = f"${total_spend:,.0f}" if total_spend >= 1000 else f"${total_spend:.2f}"
            else:
                spend_str = str(total_spend)

            lines.append(
                f"This report analyzes **Campaign {campaign_id}** over **{total_days} days** "
                f"with total spend of **{spend_str}**. "
            )
        else:
            lines.append("This report analyzes campaign performance metrics. ")

        # Key finding (first insight or derived from metrics)
        if insights:
            # Extract key message from first insight
            first_insight = insights[0]
            # Remove markdown formatting for cleaner summary
            key_finding = re.sub(r'[**#\-]', '', first_insight.split('\n')[0])
            lines.append(f"**Key Finding:** {key_finding}")
        elif metrics:
            # Derive from metrics
            ctr = metrics.get("avg_ctr", 0)
            cvr = metrics.get("avg_cvr", 0)
            roi = metrics.get("avg_roi", 0)

            if ctr > 1.5 and cvr > 3.0:
                lines.append(
                    f"The campaign is performing **strongly** with "
                    f"CTR of {ctr:.2f}% and CVR of {cvr:.2f}%, "
                    f"generating {roi:.2f}x ROI."
                )
            elif ctr < 1.0 or cvr < 2.0:
                lines.append(
                    f"The campaign is showing **below-average performance** "
                    f"with CTR of {ctr:.2f}% and CVR of {cvr:.2f}%, "
                    f"requiring optimization attention."
                )
            else:
                lines.append(
                    f"The campaign shows **average performance** "
                    f"with CTR of {ctr:.2f}% and CVR of {cvr:.2f}%."
                )

        return "\n\n".join(lines) + "\n"

    def create_metrics_section(
        self,
        metrics: Dict[str, Any],
        charts: Optional[List[str]] = None,
        prior_period: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Create the key metrics overview section.

        Includes:
        - Current period metrics
        - Comparison to prior period (if available)
        - Embedded charts
        - Performance indicators

        Args:
            metrics: Current period metrics dictionary
            charts: Optional list of chart file paths
            prior_period: Optional prior period metrics for comparison

        Returns:
            Metrics section as markdown string
        """
        lines = []
        lines.append("## Key Metrics Overview\n")

        # Create metrics table
        lines.append("### Performance Summary\n")

        if prior_period:
            # Comparison table
            lines.append("| Metric | Current Period | Prior Period | Change | Trend |")
            lines.append("|--------|----------------|--------------|--------|-------|")

            metric_definitions = {
                "impressions": ("Impressions", "{:,}", False),
                "clicks": ("Clicks", "{:,}", False),
                "conversions": ("Conversions", "{:,}", False),
                "ctr": ("CTR", "{:.2f}%", True, 100),
                "cvr": ("CVR", "{:.2f}%", True, 100),
                "cpa": ("CPA", "${:.2f}", False),
                "roi": ("ROI", "{:.2f}x", False),
                "spend": ("Total Spend", "${:,.2f}", False),
            }

            for key, (label, fmt, is_pct, mult) in metric_definitions.items():
                current = metrics.get(f"avg_{key}" if key not in ["impressions", "clicks", "conversions", "spend"] else key,
                                     metrics.get(key, metrics.get(f"total_{key}", 0)))

                prior = prior_period.get(f"avg_{key}" if key not in ["impressions", "clicks", "conversions", "spend"] else key,
                                         prior_period.get(key, prior_period.get(f"total_{key}", 0)))

                if current is None or prior is None:
                    continue

                # Format values
                if isinstance(fmt, str):
                    mult = mult if is_pct else 1
                    current_str = fmt.format(current * mult if is_pct else current)
                    prior_str = fmt.format(prior * mult if is_pct else prior)
                else:
                    current_str = str(current)
                    prior_str = str(prior)

                # Calculate change
                if prior != 0:
                    change_pct = ((current - prior) / abs(prior)) * 100
                else:
                    change_pct = 0

                change_str = f"{change_pct:+.1f}%"

                # Trend indicator
                if key in ["cpa"]:  # Lower is better
                    trend = "📉" if change_pct < -5 else "📈" if change_pct > 5 else "➡️"
                else:  # Higher is better
                    trend = "📈" if change_pct > 5 else "📉" if change_pct < -5 else "➡️"

                lines.append(f"| {label} | {current_str} | {prior_str} | {change_str} | {trend} |")
        else:
            # Simple table
            lines.append("| Metric | Value | Benchmark | Status |")
            lines.append("|--------|-------|------------|--------|")

            benchmarks = {
                "ctr": 1.2,
                "cvr": 2.5,
                "cpa": 350,
                "roi": 2.8,
            }

            metrics_to_show = [
                ("impressions", "Impressions", "{:,}", None),
                ("clicks", "Clicks", "{:,}", None),
                ("conversions", "Conversions", "{:,}", None),
                ("ctr", "CTR", "{:.2f}%", 1.2),
                ("cvr", "CVR", "{:.2f}%", 2.5),
                ("cpa", "CPA", "${:.2f}", 350),
                ("roi", "ROI", "{:.2f}x", 2.8),
                ("spend", "Total Spend", "${:,.2f}", None),
            ]

            for key, label, fmt, benchmark in metrics_to_show:
                # Get value (try avg_ prefix, then key, then total_ prefix)
                value = metrics.get(f"avg_{key}", metrics.get(key, metrics.get(f"total_{key}", 0)))

                if value is None:
                    continue

                value_str = fmt.format(value)

                # Status
                if benchmark:
                    if key == "cpa":  # Lower is better
                        status = "✅ Good" if value < benchmark else "⚠️ High" if value > benchmark * 1.2 else "➡️ OK"
                    else:  # Higher is better
                        status = "✅ Good" if value > benchmark else "⚠️ Low" if value < benchmark * 0.8 else "➡️ OK"
                    benchmark_str = f"{benchmark}"
                else:
                    status = "➡️ OK"
                    benchmark_str = "—"

                lines.append(f"| {label} | {value_str} | {benchmark_str} | {status} |")

        lines.append("\n")

        # Add charts
        if charts:
            lines.append("### Visualizations\n")
            for chart_path in charts:
                chart_name = Path(chart_path).name
                # Try to make path relative to output
                rel_path = self._make_relative_path(chart_path)
                lines.append(f"![{chart_name}]({rel_path})\n")

        return "\n".join(lines)

    def create_anomaly_section(
        self,
        anomalies: List[Dict[str, Any]],
        max_anomalies: int = 20,
    ) -> str:
        """
        Create the anomalies detected section.

        Groups anomalies by:
        - Severity (high, medium, low)
        - Metric type
        - Date

        Args:
            anomalies: List of anomaly dictionaries
            max_anomalies: Maximum number to display

        Returns:
            Anomaly section as markdown string
        """
        if not anomalies:
            return "## Anomalies Detected\n\n✅ No significant anomalies detected in the analysis period.\n"

        lines = []
        lines.append("## Anomalies Detected\n")

        # Summary
        high_count = sum(1 for a in anomalies if a.get("severity") == "high")
        medium_count = sum(1 for a in anomalies if a.get("severity") == "medium")
        low_count = sum(1 for a in anomalies if a.get("severity") == "low")

        lines.append(
            f"**Summary:** Found {len(anomalies)} anomalies "
            f"({high_count} high, {medium_count} medium, {low_count} low severity).\n"
        )

        # Group by severity
        by_severity = {"high": [], "medium": [], "low": []}
        for anomaly in anomalies:
            severity = anomaly.get("severity", "low")
            by_severity[severity].append(anomaly)

        # Display high severity first
        for severity in ["high", "medium", "low"]:
            items = by_severity[severity][:max_anomalies // 3]  # Limit per severity
            if not items:
                continue

            severity_emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}
            lines.append(f"### {severity_emoji[severity]} {severity.title()} Severity Anomalies\n")

            for anomaly in items:
                date = anomaly.get("date", "Unknown")
                metric = anomaly.get("metric_name", anomaly.get("metric", "Unknown"))
                value = anomaly.get("value", "N/A")
                score = anomaly.get("score", 0)

                lines.append(f"- **{metric}** on {date}")
                lines.append(f"  - Value: {value:.2f}")
                lines.append(f"  - Anomaly Score: {score:.2f}")

                if "direction" in anomaly:
                    lines.append(f"  - Direction: {anomaly['direction']}")
                if "magnitude_percent" in anomaly:
                    lines.append(f"  - Magnitude: {anomaly['magnitude_percent']:.1f}% deviation")

                lines.append("")

        return "\n".join(lines)

    def create_trend_section(
        self,
        trends: Dict[str, Any],
        forecast: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Create the performance trends section.

        Shows:
        - Trend direction and strength
        - Statistical significance
        - R² values
        - 7-day forecast (if available)

        Args:
            trends: Dictionary of trend analysis results by metric
            forecast: Optional forecast data

        Returns:
            Trends section as markdown string
        """
        if not trends:
            return "## Performance Trends\n\nNo trend analysis available.\n"

        lines = []
        lines.append("## Performance Trends\n")

        # Summary table
        lines.append("| Metric | Direction | Strength | R² | Significance |")
        lines.append("|--------|-----------|----------|-----|--------------|")

        for metric, trend_data in trends.items():
            direction = trend_data.get("trend_direction", "unknown")
            strength = trend_data.get("trend_strength", "unknown")
            r_squared = trend_data.get("r_squared", 0)
            p_value = trend_data.get("p_value", 1)

            # Direction emoji
            direction_emoji = {"up": "📈", "down": "📉", "flat": "➡️"}.get(direction, "❓")

            # Significance
            if p_value < 0.01:
                sig = "p < 0.01 ***"
            elif p_value < 0.05:
                sig = "p < 0.05 **"
            elif p_value < 0.1:
                sig = "p < 0.1 *"
            else:
                sig = "ns"

            lines.append(
                f"| {metric.upper()} | {direction_emoji} {direction} | {strength} | "
                f"{r_squared:.3f} | {sig} |"
            )

        lines.append("\n")

        # Detailed trend analysis
        lines.append("### Trend Details\n")

        for metric, trend_data in trends.items():
            direction = trend_data.get("trend_direction", "unknown")
            strength = trend_data.get("trend_strength", "unknown")
            interpretation = trend_data.get("interpretation", "")
            slope = trend_data.get("slope", 0)
            r_squared = trend_data.get("r_squared", 0)

            lines.append(f"#### {metric.upper()}\n")

            # Trend assessment
            if direction == "up" and strength in ["strong", "moderate"]:
                lines.append(f"✅ **Positive Trend:** {metric.upper()} is showing a {strength} upward trend.")
            elif direction == "down" and strength in ["strong", "moderate"]:
                lines.append(f"⚠️ **Declining Trend:** {metric.upper()} is showing a {strength} downward trend.")
            elif direction == "flat":
                lines.append(f"➡️ **Stable:** {metric.upper()} is relatively stable.")
            else:
                lines.append(f"📊 **{direction.title()} Trend:** {metric.upper()} is trending {direction}.")

            lines.append(f"\n- **Slope:** {slope:.4f} per day")
            lines.append(f"- **R²:** {r_squared:.3f} (model fit)")
            lines.append(f"- **Interpretation:** {interpretation}")

            # Forecast if available
            if forecast and metric in forecast:
                forecast_data = forecast[metric]
                forecast_values = forecast_data.get("forecast", [])
                if forecast_values:
                    lines.append(f"\n**7-Day Forecast:**")
                    for i, val in enumerate(forecast_values[:3], 1):
                        lines.append(f"- Day {i}: {val:.2f}")
                    if len(forecast_values) > 3:
                        lines.append(f"- ... ({len(forecast_values) - 3} more days)")

            lines.append("")

        return "\n".join(lines)

    def create_benchmark_section(
        self,
        benchmark_comparison: Dict[str, Any],
    ) -> str:
        """
        Create the benchmark comparison section.

        Compares campaign performance to industry benchmarks with:
        - Percentile rankings
        - Performance categories
        - Gap analysis

        Args:
            benchmark_comparison: Benchmark comparison dictionary

        Returns:
            Benchmark section as markdown string
        """
        if not benchmark_comparison:
            return ""

        lines = []
        lines.append("## Benchmark Comparison\n")

        # Overall performance
        overall = benchmark_comparison.get("overall_performance", "unknown")
        percentile = benchmark_comparison.get("percentile_ranking", 0)
        industry = benchmark_comparison.get("industry", "healthcare")

        lines.append(f"**Overall Performance:** {overall.upper()}")
        lines.append(f"**Industry:** {industry}")
        lines.append(f"**Percentile Ranking:** {percentile:.1f}th percentile\n")

        # Performance emoji
        if percentile >= 75:
            emoji = "🌟"
            assessment = "Top Quartile"
        elif percentile >= 50:
            emoji = "✅"
            assessment = "Above Average"
        elif percentile >= 25:
            emoji = "➡️"
            assessment = "Average"
        else:
            emoji = "⚠️"
            assessment = "Below Average"

        lines.append(f"**Assessment:** {emoji} {assessment}\n")

        # Metric-by-metric comparison
        comparisons = benchmark_comparison.get("comparisons", [])
        if comparisons:
            lines.append("### Metrics vs. Industry Benchmarks\n")
            lines.append("| Metric | Campaign | Industry Median | Difference | Percentile | Performance |")
            lines.append("|--------|----------|-----------------|------------|------------|-------------|")

            for comp in comparisons:
                metric = comp.get("metric", "Unknown")
                campaign_val = comp.get("campaign_value", 0)
                benchmark_val = comp.get("benchmark_median", 0)
                diff = comp.get("difference_from_median", 0)
                diff_pct = comp.get("difference_percent", 0)
                percentile = comp.get("percentile", 50)
                performance = comp.get("performance", "average")

                # Performance emoji
                perf_emoji = {
                    "excellent": "🌟",
                    "good": "✅",
                    "average": "➡️",
                    "below_average": "⚠️"
                }.get(performance, "❓")

                # Format difference
                diff_str = f"{diff:+.2f} ({diff_pct:+.1f}%)"

                lines.append(
                    f"| {metric} | {campaign_val:.2f} | {benchmark_val:.2f} | "
                    f"{diff_str} | {percentile:.0f}th | {perf_emoji} {performance} |"
                )

        lines.append("\n")

        return "\n".join(lines)

    def create_insights_section(
        self,
        insights: List[str],
        max_insights: int = 10,
    ) -> str:
        """
        Create the insights section.

        Formats insights as a numbered list with proper markdown.

        Args:
            insights: List of insight strings
            max_insights: Maximum number to display

        Returns:
            Insights section as markdown string
        """
        if not insights:
            return "## Key Insights\n\nNo insights generated.\n"

        lines = []
        lines.append("## Key Insights\n")

        for i, insight in enumerate(insights[:max_insights], 1):
            # Clean up markdown if it has embedded numbering
            cleaned = re.sub(r'^[\d\-\*\.]+\s*', '', insight.strip())
            lines.append(f"{i}. {cleaned}")

        lines.append("")

        return "\n".join(lines)

    def create_recommendations_section(
        self,
        recommendations: List[str],
        prioritize: bool = True,
    ) -> str:
        """
        Create the recommendations section.

        Prioritizes and formats recommendations as actionable items.

        Args:
            recommendations: List of recommendation strings
            prioritize: Whether to prioritize by impact keywords

        Returns:
            Recommendations section as markdown string
        """
        if not recommendations:
            return "## Recommendations\n\nNo specific recommendations at this time.\n"

        lines = []
        lines.append("## Recommendations\n")

        # Prioritize if enabled
        if prioritize:
            prioritized = self._prioritize_recommendations(recommendations)
        else:
            prioritized = recommendations

        for i, rec in enumerate(prioritized, 1):
            # Clean up any existing formatting
            cleaned = re.sub(r'^[\d\-\*\.]+\s*\*\*', '', rec.strip())
            cleaned = re.sub(r'^[\d\-\*\.]+\s*', '', cleaned)

            # Add priority indicator
            priority = self._get_priority_level(cleaned)
            priority_emoji = {
                "high": "🔴",
                "medium": "🟡",
                "low": "🟢"
            }.get(priority, "⚪")

            lines.append(f"{i}. {priority_emoji} {cleaned}")

        lines.append("")

        return "\n".join(lines)

    def create_appendix_section(
        self,
        raw_data: Optional[Dict[str, Any]] = None,
        metrics: Optional[Dict[str, Any]] = None,
    ) -> str:
        """
        Create the appendix with raw data tables.

        Args:
            raw_data: Raw campaign data dictionary
            metrics: Metrics summary dictionary

        Returns:
            Appendix section as markdown string
        """
        if not self.include_raw_data:
            return ""

        lines = []
        lines.append("## Appendix: Raw Data\n")

        # Metrics detail table
        if metrics:
            lines.append("### Detailed Metrics\n")

            for key, value in metrics.items():
                if isinstance(value, (int, float)):
                    lines.append(f"- **{key}:** {value}")
                elif isinstance(value, dict):
                    lines.append(f"\n#### {key}\n")
                    for sub_key, sub_value in value.items():
                        lines.append(f"- {sub_key}: {sub_value}")
                elif isinstance(value, list):
                    lines.append(f"\n#### {key}\n")
                    lines.append(f"({len(value)} items)")

        # Raw data sample
        if raw_data and "data" in raw_data:
            data_list = raw_data["data"]
            if data_list and len(data_list) > 0:
                lines.append("\n### Data Sample\n")

                # Convert to DataFrame for nice table
                df = pd.DataFrame(data_list[:10])  # First 10 rows
                lines.append(df.to_markdown(index=False))
                lines.append(f"\n*Showing first 10 of {len(data_list)} records*\n")

        return "\n".join(lines)

    # =========================================================================
    # FULL REPORT GENERATION
    # =========================================================================

    def generate_full_report(
        self,
        state: Dict[str, Any],
        title: Optional[str] = None,
        include_sections: Optional[List[str]] = None,
    ) -> str:
        """
        Generate a complete markdown report from agent state.

        Sections (in order):
        1. Header (title, metadata)
        2. Executive Summary
        3. Key Metrics Overview
        4. Performance Trends
        5. Anomalies Detected
        6. Benchmark Comparison
        7. Insights
        8. Recommendations
        9. Appendix

        Args:
            state: Agent state dictionary with all analysis results
            title: Optional report title
            include_sections: Optional list of sections to include (default: all)

        Returns:
            Complete markdown report as string
        """
        lines = []

        # Default: include all sections
        all_sections = [
            "executive_summary",
            "metrics",
            "trends",
            "anomalies",
            "benchmark",
            "insights",
            "recommendations",
            "appendix",
        ]

        if include_sections is None:
            include_sections = all_sections

        # Header
        lines.extend(self._create_header(state, title))
        lines.append("\n---\n")

        # Executive Summary
        if "executive_summary" in include_sections:
            lines.append(self.create_executive_summary(
                insights=state.get("insights", []),
                metrics=state.get("metrics_summary")
            ))
            lines.append("\n")

        # Key Metrics
        if "metrics" in include_sections and state.get("metrics_summary"):
            lines.append(self.create_metrics_section(
                metrics=state.get("metrics_summary", {}),
                charts=state.get("charts", [])
            ))
            lines.append("\n")

        # Performance Trends
        if "trends" in include_sections and state.get("trends"):
            lines.append(self.create_trend_section(
                trends=state.get("trends", {}),
            ))
            lines.append("\n")

        # Anomalies
        if "anomalies" in include_sections and state.get("anomalies"):
            lines.append(self.create_anomaly_section(
                anomalies=state.get("anomalies", []),
            ))
            lines.append("\n")

        # Benchmark Comparison
        if "benchmark" in include_sections and state.get("benchmark_comparison"):
            lines.append(self.create_benchmark_section(
                benchmark_comparison=state.get("benchmark_comparison", {}),
            ))
            lines.append("\n")

        # Insights
        if "insights" in include_sections and state.get("insights"):
            lines.append(self.create_insights_section(
                insights=state.get("insights", []),
            ))
            lines.append("\n")

        # Recommendations
        if "recommendations" in include_sections and state.get("recommendations"):
            lines.append(self.create_recommendations_section(
                recommendations=state.get("recommendations", []),
            ))
            lines.append("\n")

        # Appendix
        if "appendix" in include_sections:
            lines.append(self.create_appendix_section(
                raw_data=state.get("raw_data"),
                metrics=state.get("metrics_summary"),
            ))
            lines.append("\n")

        # Footer
        lines.extend(self._create_footer())

        return "\n".join(lines)

    # =========================================================================
    # EXPORT FUNCTIONS
    # =========================================================================

    def export_to_html(
        self,
        markdown: str,
        output_filename: Optional[str] = None,
    ) -> str:
        """
        Convert markdown report to HTML.

        Args:
            markdown: Markdown content
            output_filename: Optional output filename

        Returns:
            Path to generated HTML file
        """
        if not MARKDOWN_AVAILABLE:
            raise ImportError(
                "markdown library is required for HTML export. "
                "Install with: pip install markdown"
            )

        # Convert markdown to HTML
        md = markdown.Markdown(
            extensions=[tables, fenced_code, nl2br]
        )
        html_content = md.convert(markdown)

        # Wrap in HTML template
        html_template = self._get_html_template()
        full_html = html_template.format(
            title="AdInsights Report",
            content=html_content,
            css=self._css,
        )

        # Generate filename
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"report_{timestamp}.html"

        output_path = self.output_dir / output_filename

        # Write file
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(full_html)

        return str(output_path)

    def export_to_pdf(
        self,
        markdown: str,
        output_filename: Optional[str] = None,
    ) -> str:
        """
        Convert markdown report to PDF.

        Note: Requires weasyprint. Install with: pip install weasyprint

        Args:
            markdown: Markdown content
            output_filename: Optional output filename

        Returns:
            Path to generated PDF file
        """
        if not WEASYPRINT_AVAILABLE:
            raise ImportError(
                "weasyprint library is required for PDF export. "
                "Install with: pip install weasyprint"
            )

        # First convert to HTML
        html_path = self.export_to_html(markdown)

        # Convert HTML to PDF
        output_path = html_path.replace(".html", ".pdf")
        if output_filename:
            output_path = str(self.output_dir / output_filename)

        HTML(html_path).write_pdf(output_path)

        return output_path

    def export_to_markdown(
        self,
        markdown: str,
        output_filename: Optional[str] = None,
    ) -> str:
        """
        Save markdown report to file.

        Args:
            markdown: Markdown content
            output_filename: Optional output filename

        Returns:
            Path to generated markdown file
        """
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"report_{timestamp}.md"

        output_path = self.output_dir / output_filename

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(markdown)

        return str(output_path)

    def export_to_excel(
        self,
        state: Dict[str, Any],
        output_filename: Optional[str] = None,
    ) -> str:
        """
        Export report data to Excel file with multiple sheets.

        Args:
            state: Agent state dictionary with all analysis results
            output_filename: Optional output filename

        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl library is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        # Generate filename
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"report_{timestamp}.xlsx"

        output_path = self.output_dir / output_filename

        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Add sheets for each section
        self._add_summary_sheet(wb, state)
        self._add_metrics_sheet(wb, state)
        self._add_trends_sheet(wb, state)
        self._add_anomalies_sheet(wb, state)
        self._add_insights_sheet(wb, state)

        # Save workbook
        wb.save(output_path)

        return str(output_path)

    def export_all(
        self,
        markdown: str,
        base_filename: Optional[str] = None,
        state: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, str]:
        """
        Export report to all available formats.

        Args:
            markdown: Markdown content
            base_filename: Base filename (without extension)
            state: Agent state for Excel export (optional)

        Returns:
            Dictionary with format -> filepath mappings
        """
        results = {}

        # Markdown (always available)
        if base_filename:
            md_file = f"{base_filename}.md"
        else:
            md_file = None
        results["markdown"] = self.export_to_markdown(markdown, md_file)

        # HTML
        if MARKDOWN_AVAILABLE:
            if base_filename:
                html_file = f"{base_filename}.html"
            else:
                html_file = None
            results["html"] = self.export_to_html(markdown, html_file)

        # PDF
        if WEASYPRINT_AVAILABLE:
            if base_filename:
                pdf_file = f"{base_filename}.pdf"
            else:
                pdf_file = None
            results["pdf"] = self.export_to_pdf(markdown, pdf_file)

        # Excel
        if OPENPYXL_AVAILABLE and state:
            if base_filename:
                excel_file = f"{base_filename}.xlsx"
            else:
                excel_file = None
            results["excel"] = self.export_to_excel(state, excel_file)

        return results

    # =========================================================================
    # HELPER METHODS
    # =========================================================================

    def _create_header(
        self,
        state: Dict[str, Any],
        title: Optional[str] = None,
    ) -> List[str]:
        """Create report header."""
        lines = []

        # Title
        if title is None:
            campaign_id = state.get("campaign_id", "Unknown")
            title = f"Campaign Analysis Report - {campaign_id}"

        lines.append(f"# {title}\n")

        # Metadata
        lines.append("**Generated by:** AdInsights Analytics")
        lines.append(f"**Company:** {self.company_name}")

        if self.include_timestamps:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            lines.append(f"**Generated:** {timestamp}")

        # Campaign details
        if state.get("campaign_id"):
            lines.append(f"**Campaign ID:** {state['campaign_id']}")

        if state.get("date_range"):
            start, end = state["date_range"]
            lines.append(f"**Analysis Period:** {start} to {end}")

        lines.append("")

        return lines

    def _create_footer(self) -> List[str]:
        """Create report footer."""
        lines = []
        lines.append("\n---\n")
        lines.append(
            f"*This report was automatically generated by "
            f"**{self.company_name}** on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.*\n"
        )
        lines.append(
            "*For questions about this report, contact your analytics team.*"
        )
        return lines

    def _prioritize_recommendations(
        self,
        recommendations: List[str],
    ) -> List[str]:
        """Prioritize recommendations by impact keywords."""
        priority_scores = {
            "high": [
                "critical", "urgent", "immediately", "asap", "severe",
                "significant", "major", "important", "priority",
                "address", "fix", "resolve"
            ],
            "medium": [
                "consider", "review", "optimize", "improve", "enhance",
                "adjust", "refine", "update", "modify"
            ],
            "low": [
                "monitor", "watch", "track", "observe", "note",
                "optional", "may", "might", "could"
            ]
        }

        scored = []
        for rec in recommendations:
            score = 0
            rec_lower = rec.lower()

            for word in priority_scores["high"]:
                if word in rec_lower:
                    score += 3
            for word in priority_scores["medium"]:
                if word in rec_lower:
                    score += 2
            for word in priority_scores["low"]:
                if word in rec_lower:
                    score += 1

            scored.append((score, rec))

        # Sort by score descending
        scored.sort(key=lambda x: x[0], reverse=True)

        return [rec for score, rec in scored]

    def _get_priority_level(self, text: str) -> str:
        """Get priority level from text."""
        text_lower = text.lower()

        high_words = ["critical", "urgent", "immediately", "severe", "significant"]
        medium_words = ["consider", "review", "optimize", "improve", "adjust"]

        if any(word in text_lower for word in high_words):
            return "high"
        elif any(word in text_lower for word in medium_words):
            return "medium"
        else:
            return "low"

    def _make_relative_path(self, file_path: str) -> str:
        """Make a file path relative to output directory."""
        try:
            path = Path(file_path)
            if path.is_absolute():
                return str(path.relative_to(self.output_dir.parent))
            return file_path
        except:
            return file_path

    def _get_html_template(self) -> str:
        """Get HTML template for reports."""
        return """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        {css}
    </style>
</head>
<body>
    <div class="container">
        {content}
    </div>
</body>
</html>"""

    def _get_default_css(self) -> str:
        """Get default CSS for HTML/PDF styling."""
        return """
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 900px;
            margin: 0 auto;
            padding: 20px;
            background: #f5f5f5;
        }

        .container {
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }

        h1, h2, h3, h4 {
            color: #2c3e50;
            margin-top: 24px;
            margin-bottom: 16px;
        }

        h1 {
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }

        h2 {
            border-bottom: 1px solid #eee;
            padding-bottom: 8px;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }

        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }

        th {
            background-color: #3498db;
            color: white;
            font-weight: 600;
        }

        tr:hover {
            background-color: #f5f5f5;
        }

        code {
            background-color: #f4f4f4;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Monaco', 'Courier New', monospace;
        }

        pre {
            background-color: #f4f4f4;
            padding: 16px;
            border-radius: 4px;
            overflow-x: auto;
        }

        blockquote {
            border-left: 4px solid #3498db;
            padding-left: 16px;
            margin: 20px 0;
            color: #666;
        }

        img {
            max-width: 100%;
            height: auto;
            border-radius: 4px;
            margin: 20px 0;
        }

        .emoji {
            font-size: 1.2em;
        }
        """

    # =========================================================================
    # EXCEL HELPER METHODS
    # =========================================================================

    def _add_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        state: Dict[str, Any],
    ) -> None:
        """Add executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)

        # Title
        ws["A1"] = f"AdInsights Analytics Report"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Campaign: {state.get('campaign_id', 'N/A')}"
        ws["A2"].font = Font(size=12, bold=True)

        if state.get("date_range"):
            start, end = state["date_range"]
            ws["A3"] = f"Period: {start} to {end}"
        ws["A3"].font = Font(size=10)

        ws["A5"] = "Key Metrics Summary"
        ws["A5"].font = Font(size=12, bold=True)

        # Metrics table
        metrics = state.get("metrics_summary", {})
        row = 6
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row += 1
        metric_labels = {
            "avg_ctr": "Average CTR (%)",
            "avg_cvr": "Average CVR (%)",
            "avg_cpa": "Average CPA ($)",
            "avg_roi": "Average ROI",
            "total_impressions": "Total Impressions",
            "total_clicks": "Total Clicks",
            "total_conversions": "Total Conversions",
            "total_spend": "Total Spend ($)",
        }

        for key, label in metric_labels.items():
            if key in metrics:
                ws.cell(row, 1, label)
                ws.cell(row, 2, metrics[key])
                row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _add_metrics_sheet(
        self,
        wb: openpyxl.Workbook,
        state: Dict[str, Any],
    ) -> None:
        """Add detailed metrics sheet."""
        ws = wb.create_sheet("Detailed Metrics")

        ws["A1"] = "Detailed Campaign Metrics"
        ws["A1"].font = Font(size=14, bold=True)

        metrics = state.get("metrics_summary", {})
        row = 3

        for key, value in metrics.items():
            ws.cell(row, 1, key.replace("_", " ").title())
            ws.cell(row, 2, value)
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _add_trends_sheet(
        self,
        wb: openpyxl.Workbook,
        state: Dict[str, Any],
    ) -> None:
        """Add trends analysis sheet."""
        ws = wb.create_sheet("Trends Analysis")

        ws["A1"] = "Trend Analysis"
        ws["A1"].font = Font(size=14, bold=True)

        trends = state.get("trends", {})
        if not trends:
            ws["A3"] = "No trend data available"
            return

        # Header row
        headers = ["Metric", "Direction", "Strength", "R²", "P-Value", "Slope", "Interpretation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        for metric, trend_data in trends.items():
            ws.cell(row, 1, metric.upper())
            ws.cell(row, 2, trend_data.get("trend_direction", "N/A"))
            ws.cell(row, 3, trend_data.get("trend_strength", "N/A"))
            ws.cell(row, 4, trend_data.get("r_squared", 0))
            ws.cell(row, 5, trend_data.get("p_value", 0))
            ws.cell(row, 6, trend_data.get("slope", 0))
            ws.cell(row, 7, trend_data.get("interpretation", ""))
            row += 1

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def _add_anomalies_sheet(
        self,
        wb: openpyxl.Workbook,
        state: Dict[str, Any],
    ) -> None:
        """Add anomalies sheet."""
        ws = wb.create_sheet("Anomalies")

        ws["A1"] = "Detected Anomalies"
        ws["A1"].font = Font(size=14, bold=True)

        anomalies = state.get("anomalies", [])
        if not anomalies:
            ws["A3"] = "No anomalies detected"
            return

        # Header row
        headers = ["Date", "Metric", "Value", "Severity", "Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        for anomaly in anomalies:
            ws.cell(row, 1, anomaly.get("date", "N/A"))
            ws.cell(row, 2, anomaly.get("metric_name", "N/A"))
            ws.cell(row, 3, anomaly.get("value", 0))
            ws.cell(row, 4, anomaly.get("severity", "N/A"))
            ws.cell(row, 5, anomaly.get("score", 0))

            # Color code severity
            severity = anomaly.get("severity", "")
            if severity == "high":
                for col in range(1, 6):
                    ws.cell(row, col).fill = PatternFill(
                        start_color="ffebee", end_color="ffebee", fill_type="solid"
                    )
            elif severity == "medium":
                for col in range(1, 6):
                    ws.cell(row, col).fill = PatternFill(
                        start_color="fff3e0", end_color="fff3e0", fill_type="solid"
                    )

            row += 1

        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def _add_insights_sheet(
        self,
        wb: openpyxl.Workbook,
        state: Dict[str, Any],
    ) -> None:
        """Add insights and recommendations sheet."""
        ws = wb.create_sheet("Insights & Recommendations")

        ws["A1"] = "Insights and Recommendations"
        ws["A1"].font = Font(size=14, bold=True)

        row = 3

        # Add insights
        insights = state.get("insights", [])
        if insights:
            ws["A3"] = "Key Insights"
            ws["A3"].font = Font(size=12, bold=True)
            row = 4
            for i, insight in enumerate(insights, 1):
                ws.cell(row, 1, f"{i}. {insight}")
                row += 2

        # Add recommendations
        recommendations = state.get("recommendations", [])
        if recommendations:
            if row < 5:
                row = 5
            ws[f"A{row}"] = "Recommendations"
            ws[f"A{row}"].font = Font(size=12, bold=True)
            row += 1
            for i, rec in enumerate(recommendations, 1):
                ws.cell(row, 1, f"{i}. {rec}")
                row += 2

        # Add benchmark comparison
        benchmark = state.get("benchmark_comparison", {})
        if benchmark:
            row += 1
            ws[f"A{row}"] = "Benchmark Comparison"
            ws[f"A{row}"].font = Font(size=12, bold=True)
            row += 1

            comparisons = benchmark.get("comparisons", [])
            if comparisons:
                # Header
                headers = ["Metric", "Campaign", "Benchmark", "Performance", "Percentile"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row, col, header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="27ae60", end_color="27ae60", fill_type="solid"
                    )
                row += 1

                for comp in comparisons:
                    ws.cell(row, 1, comp.get("metric", "N/A"))
                    ws.cell(row, 2, comp.get("campaign_value", 0))
                    ws.cell(row, 3, comp.get("benchmark_median", 0))
                    ws.cell(row, 4, comp.get("performance", "N/A"))
                    ws.cell(row, 5, comp.get("percentile", 0))
                    row += 1

        ws.column_dimensions["A"].width = 50


# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def quick_report(
    state: Dict[str, Any],
    output_format: str = "markdown",
    output_dir: str = "./data/reports",
) -> str:
    """
    Generate a report from agent state.

    Args:
        state: Agent state dictionary
        output_format: Format to export (markdown, html, pdf, all)
        output_dir: Output directory

    Returns:
        Path to generated report(s)
    """
    generator = ReportGenerator(output_dir=output_dir)

    # Generate markdown
    markdown = generator.generate_full_report(state)

    # Export
    if output_format == "markdown":
        return generator.export_to_markdown(markdown)
    elif output_format == "html":
        return generator.export_to_html(markdown)
    elif output_format == "pdf":
        return generator.export_to_pdf(markdown)
    elif output_format == "all":
        results = generator.export_all(markdown)
        return "\n".join([f"{fmt}: {path}" for fmt, path in results.items()])
    else:
        raise ValueError(f"Unknown output format: {output_format}")


# =============================================================================
# MAIN DEMONSTRATION
# =============================================================================

if __name__ == "__main__":
    print("Report Generator for AdInsights-Agent")
    print("=" * 60)

    # Create sample state
    sample_state = {
        "campaign_id": "CAMP-DEMO-001",
        "date_range": ("2024-01-01", "2024-01-30"),
        "metrics_summary": {
            "avg_ctr": 1.35,
            "avg_cvr": 2.78,
            "avg_cpa": 285.50,
            "avg_roi": 3.12,
            "total_impressions": 425000,
            "total_clicks": 5742,
            "total_conversions": 159,
            "total_spend": 45400.00,
        },
        "trends": {
            "ctr": {
                "trend_direction": "up",
                "trend_strength": "moderate",
                "r_squared": 0.65,
                "p_value": 0.02,
                "slope": 0.015,
                "interpretation": "CTR is trending upward with moderate strength.",
            },
            "cvr": {
                "trend_direction": "down",
                "trend_strength": "weak",
                "r_squared": 0.35,
                "p_value": 0.12,
                "slope": -0.008,
                "interpretation": "CVR shows a slight downward trend but not statistically significant.",
            }
        },
        "anomalies": [
            {
                "date": "2024-01-15",
                "metric_name": "CTR",
                "value": 2.8,
                "severity": "high",
                "score": 3.2,
            },
            {
                "date": "2024-01-22",
                "metric_name": "CVR",
                "value": 1.2,
                "severity": "medium",
                "score": 2.6,
            }
        ],
        "benchmark_comparison": {
            "overall_performance": "good",
            "percentile_ranking": 68.5,
            "industry": "healthcare_pharma",
            "comparisons": [
                {
                    "metric": "CTR",
                    "campaign_value": 1.35,
                    "benchmark_median": 1.2,
                    "performance": "good",
                    "percentile": 65,
                },
                {
                    "metric": "CVR",
                    "campaign_value": 2.78,
                    "benchmark_median": 2.5,
                    "performance": "average",
                    "percentile": 55,
                }
            ]
        },
        "insights": [
            "Campaign CTR is performing 12.5% above industry median.",
            "CVR has declined 8% over the analysis period.",
            "2 high-severity anomalies detected on January 15th.",
        ],
        "recommendations": [
            "Investigate the CTR spike on January 15th for potential fraud or technical issues.",
            "Optimize landing pages to address declining CVR trend.",
            "Consider allocating more budget to high-performing ad creatives.",
        ],
        "charts": [],
        "raw_data": None,
    }

    # Generate report
    generator = ReportGenerator()

    print("\nGenerating report...")
    markdown_report = generator.generate_full_report(sample_state)

    print("\n" + "=" * 60)
    print(markdown_report)
    print("=" * 60)

    # Try to export
    print("\nExporting report...")

    try:
        md_path = generator.export_to_markdown(markdown_report, "demo_report.md")
        print(f"✅ Markdown: {md_path}")
    except Exception as e:
        print(f"❌ Markdown export failed: {e}")

    try:
        html_path = generator.export_to_html(markdown_report, "demo_report.html")
        print(f"✅ HTML: {html_path}")
    except Exception as e:
        print(f"⚠️  HTML export not available: {e}")

    try:
        pdf_path = generator.export_to_pdf(markdown_report, "demo_report.pdf")
        print(f"✅ PDF: {pdf_path}")
    except Exception as e:
        print(f"⚠️  PDF export not available: {e}")

    print("\n" + "=" * 60)
    print("Report generation complete!")
